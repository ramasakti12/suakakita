from flask import Blueprint, jsonify, send_file
from flask_login import login_required
from src.models.melon_pinter import MelonPinterData
from src.models.kangen_azi import KangenAziData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime, timedelta
import os
import tempfile

export_bp = Blueprint('export', __name__)

def create_excel_file():
    """Create Excel file with both Melon Pinter and Kangen Azi data"""
    # Get data from last year
    one_year_ago = datetime.utcnow() - timedelta(days=365)
    
    melon_data = MelonPinterData.query.filter(
        MelonPinterData.timestamp >= one_year_ago
    ).order_by(MelonPinterData.timestamp.desc()).all()
    
    kangen_data = KangenAziData.query.filter(
        KangenAziData.timestamp >= one_year_ago
    ).order_by(KangenAziData.timestamp.desc()).all()
    
    # Create workbook
    wb = Workbook()
    
    # Create Melon Pinter sheet
    ws_melon = wb.active
    ws_melon.title = "Melon Pinter Data"
    
    # Melon Pinter headers
    melon_headers = [
        'Timestamp', 'User ID', 'Nama', 'Jenis Kelamin', 'Usia', 
        'Ekspresi', 'Keluhan Fisik', 'Heart Rate (bpm)', 
        'Respiration Rate (/min)', 'Suhu (°C)', 'Peringatan'
    ]
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(melon_headers, 1):
        cell = ws_melon.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add Melon Pinter data
    for row, entry in enumerate(melon_data, 2):
        ws_melon.cell(row=row, column=1, value=entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws_melon.cell(row=row, column=2, value=entry.user_id)
        ws_melon.cell(row=row, column=3, value=entry.name)
        ws_melon.cell(row=row, column=4, value=entry.gender)
        ws_melon.cell(row=row, column=5, value=entry.age)
        ws_melon.cell(row=row, column=6, value=entry.expression)
        ws_melon.cell(row=row, column=7, value=entry.complaints or '')
        ws_melon.cell(row=row, column=8, value=entry.heart_rate)
        ws_melon.cell(row=row, column=9, value=entry.respiration_rate)
        ws_melon.cell(row=row, column=10, value=entry.temperature)
        ws_melon.cell(row=row, column=11, value=entry.warnings or '')
    
    # Create Kangen Azi sheet
    ws_kangen = wb.create_sheet("Kangen Azi Data")
    
    # Kangen Azi headers
    kangen_headers = [
        'Timestamp', 'User ID', 'Berat (kg)', 'Tinggi (cm)', 'Usia', 
        'Olahraga', 'BMI', 'Kategori BMI', 'Status', 'Rekomendasi Diet'
    ]
    
    for col, header in enumerate(kangen_headers, 1):
        cell = ws_kangen.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add Kangen Azi data
    for row, entry in enumerate(kangen_data, 2):
        ws_kangen.cell(row=row, column=1, value=entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
        ws_kangen.cell(row=row, column=2, value=entry.user_id)
        ws_kangen.cell(row=row, column=3, value=entry.weight)
        ws_kangen.cell(row=row, column=4, value=entry.height)
        ws_kangen.cell(row=row, column=5, value=entry.age)
        ws_kangen.cell(row=row, column=6, value=entry.sport)
        ws_kangen.cell(row=row, column=7, value=entry.bmi)
        ws_kangen.cell(row=row, column=8, value=entry.bmi_category)
        ws_kangen.cell(row=row, column=9, value=entry.bmi_status)
        ws_kangen.cell(row=row, column=10, value=entry.diet_recommendations)
    
    # Auto-adjust column widths
    for ws in [ws_melon, ws_kangen]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    return wb

@export_bp.route('/download', methods=['GET'])
@login_required
def download_excel():
    """Download Excel file with all data"""
    try:
        wb = create_excel_file()
        
        # Create temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Generate filename with current date
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f'Student_Health_Data_{current_date}.xlsx'
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/auto-export', methods=['POST'])
@login_required
def auto_export():
    """Auto export to downloads folder"""
    try:
        wb = create_excel_file()
        
        # Create downloads directory if it doesn't exist
        downloads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'downloads')
        os.makedirs(downloads_dir, exist_ok=True)
        
        # Generate filename with current date
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f'Student_Health_Data_{current_date}.xlsx'
        filepath = os.path.join(downloads_dir, filename)
        
        wb.save(filepath)
        
        return jsonify({
            'message': 'File exported successfully',
            'filename': filename,
            'filepath': filepath
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@export_bp.route('/summary', methods=['GET'])
@login_required
def get_export_summary():
    """Get summary of data available for export"""
    try:
        one_year_ago = datetime.utcnow() - timedelta(days=365)
        
        melon_count = MelonPinterData.query.filter(
            MelonPinterData.timestamp >= one_year_ago
        ).count()
        
        kangen_count = KangenAziData.query.filter(
            KangenAziData.timestamp >= one_year_ago
        ).count()
        
        return jsonify({
            'melon_pinter_records': melon_count,
            'kangen_azi_records': kangen_count,
            'total_records': melon_count + kangen_count,
            'data_range': '1 year'
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

