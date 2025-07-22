from apscheduler.schedulers.background import BackgroundScheduler
from apscheduler.triggers.cron import CronTrigger
from datetime import datetime
import os
import logging
from src.models.melon_pinter import MelonPinterData
from src.models.kangen_azi import KangenAziData
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def create_monthly_export():
    """Create monthly export of all data"""
    try:
        logger.info("Starting monthly export...")
        
        # Get all data from last month
        from datetime import datetime, timedelta
        one_month_ago = datetime.utcnow() - timedelta(days=30)
        
        melon_data = MelonPinterData.query.filter(
            MelonPinterData.timestamp >= one_month_ago
        ).order_by(MelonPinterData.timestamp.desc()).all()
        
        kangen_data = KangenAziData.query.filter(
            KangenAziData.timestamp >= one_month_ago
        ).order_by(KangenAziData.timestamp.desc()).all()
        
        # Create workbook
        wb = Workbook()
        
        # Create Melon Pinter sheet
        ws_melon = wb.active
        ws_melon.title = "Melon Pinter Data"
        
        # Melon Pinter headers
        melon_headers = [
            'Timestamp', 'User ID', 'Nama', 'Jenis Kelamin', 'Usia', 
            'Ekspresi', 'Keluhan Fisik', 'Heart Rate (bpm)', 
            'Respiration Rate (/min)', 'Suhu (°C)', 'Peringatan'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(melon_headers, 1):
            cell = ws_melon.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add Melon Pinter data
        for row, entry in enumerate(melon_data, 2):
            ws_melon.cell(row=row, column=1, value=entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws_melon.cell(row=row, column=2, value=entry.user_id)
            ws_melon.cell(row=row, column=3, value=entry.name)
            ws_melon.cell(row=row, column=4, value=entry.gender)
            ws_melon.cell(row=row, column=5, value=entry.age)
            ws_melon.cell(row=row, column=6, value=entry.expression)
            ws_melon.cell(row=row, column=7, value=entry.complaints or '')
            ws_melon.cell(row=row, column=8, value=entry.heart_rate)
            ws_melon.cell(row=row, column=9, value=entry.respiration_rate)
            ws_melon.cell(row=row, column=10, value=entry.temperature)
            ws_melon.cell(row=row, column=11, value=entry.warnings or '')
        
        # Create Kangen Azi sheet
        ws_kangen = wb.create_sheet("Kangen Azi Data")
        
        # Kangen Azi headers
        kangen_headers = [
            'Timestamp', 'User ID', 'Berat (kg)', 'Tinggi (cm)', 'Usia', 
            'Olahraga', 'BMI', 'Kategori BMI', 'Status', 'Rekomendasi Diet'
        ]
        
        for col, header in enumerate(kangen_headers, 1):
            cell = ws_kangen.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add Kangen Azi data
        for row, entry in enumerate(kangen_data, 2):
            ws_kangen.cell(row=row, column=1, value=entry.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
            ws_kangen.cell(row=row, column=2, value=entry.user_id)
            ws_kangen.cell(row=row, column=3, value=entry.weight)
            ws_kangen.cell(row=row, column=4, value=entry.height)
            ws_kangen.cell(row=row, column=5, value=entry.age)
            ws_kangen.cell(row=row, column=6, value=entry.sport)
            ws_kangen.cell(row=row, column=7, value=entry.bmi)
            ws_kangen.cell(row=row, column=8, value=entry.bmi_category)
            ws_kangen.cell(row=row, column=9, value=entry.bmi_status)
            ws_kangen.cell(row=row, column=10, value=entry.diet_recommendations)
        
        # Auto-adjust column widths
        for ws in [ws_melon, ws_kangen]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create downloads directory if it doesn't exist
        downloads_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.dirname(__file__))), 'downloads')
        os.makedirs(downloads_dir, exist_ok=True)
        
        # Generate filename with current date
        current_date = datetime.now().strftime('%Y-%m-%d')
        filename = f'Monthly_Export_Student_Health_Data_{current_date}.xlsx'
        filepath = os.path.join(downloads_dir, filename)
        
        wb.save(filepath)
        
        logger.info(f"Monthly export completed successfully: {filename}")
        logger.info(f"Exported {len(melon_data)} Melon Pinter records and {len(kangen_data)} Kangen Azi records")
        
        return True
        
    except Exception as e:
        logger.error(f"Monthly export failed: {str(e)}")
        return False

def cleanup_old_data():
    """Clean up data older than 1 year"""
    try:
        logger.info("Starting data cleanup...")
        
        from datetime import datetime, timedelta
        one_year_ago = datetime.utcnow() - timedelta(days=365)
        
        # Count records to be deleted
        melon_count = MelonPinterData.query.filter(
            MelonPinterData.timestamp < one_year_ago
        ).count()
        
        kangen_count = KangenAziData.query.filter(
            KangenAziData.timestamp < one_year_ago
        ).count()
        
        if melon_count > 0 or kangen_count > 0:
            # Delete old records
            MelonPinterData.query.filter(
                MelonPinterData.timestamp < one_year_ago
            ).delete()
            
            KangenAziData.query.filter(
                KangenAziData.timestamp < one_year_ago
            ).delete()
            
            # Commit changes
            from src.models.user import db
            db.session.commit()
            
            logger.info(f"Cleanup completed: Deleted {melon_count} Melon Pinter records and {kangen_count} Kangen Azi records")
        else:
            logger.info("No old data to cleanup")
        
        return True
        
    except Exception as e:
        logger.error(f"Data cleanup failed: {str(e)}")
        return False

class HealthMonitorScheduler:
    def __init__(self):
        self.scheduler = BackgroundScheduler()
        self.scheduler.start()
        logger.info("Health Monitor Scheduler started")
    
    def setup_jobs(self):
        """Setup scheduled jobs"""
        try:
            # Monthly export on the 1st day of each month at 2:00 AM
            self.scheduler.add_job(
                func=create_monthly_export,
                trigger=CronTrigger(day=1, hour=2, minute=0),
                id='monthly_export',
                name='Monthly Data Export',
                replace_existing=True
            )
            
            # Data cleanup every 3 months on the 15th at 3:00 AM
            self.scheduler.add_job(
                func=cleanup_old_data,
                trigger=CronTrigger(month='*/3', day=15, hour=3, minute=0),
                id='data_cleanup',
                name='Data Cleanup (1 year retention)',
                replace_existing=True
            )
            
            logger.info("Scheduled jobs configured:")
            logger.info("- Monthly export: 1st day of each month at 2:00 AM")
            logger.info("- Data cleanup: Every 3 months on 15th at 3:00 AM")
            
        except Exception as e:
            logger.error(f"Failed to setup scheduled jobs: {str(e)}")
    
    def shutdown(self):
        """Shutdown the scheduler"""
        self.scheduler.shutdown()
        logger.info("Health Monitor Scheduler shutdown")
    
    def get_jobs(self):
        """Get list of scheduled jobs"""
        return self.scheduler.get_jobs()

# Global scheduler instance
scheduler_instance = None

def init_scheduler():
    """Initialize the scheduler"""
    global scheduler_instance
    if scheduler_instance is None:
        scheduler_instance = HealthMonitorScheduler()
        scheduler_instance.setup_jobs()
    return scheduler_instance

def get_scheduler():
    """Get the scheduler instance"""
    return scheduler_instance

